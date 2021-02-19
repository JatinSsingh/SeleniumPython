import openpyxl

book = openpyxl.load_workbook("C:\\Users\\Jatin Singh\\Downloads\\dataA (1).xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=3)
print(cell.value)

sheet.cell(row=2, column=2).value = "Jatin Ssingh"
print(sheet.cell(row=2, column=2).value)

print(sheet.max_row)
print(sheet.max_column)
print(sheet['B2'].value)

